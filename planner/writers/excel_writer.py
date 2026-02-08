"""
Excel output generation.

Responsibilities:
- Create workbooks and worksheets
- Apply formatting and styling
- Render planning rows into Excel layout
- Handle metadata and presentation details

No planning or date logic should live here.
If Excel looks wrong but logic is right, debug here.
"""


from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

import socket

class ExcelWriter:

    def __init__(self):
        self.wb=Workbook()


    def write_subject_sheet(self, subject_name: str, rows: list[dict]):
        ws = self.wb.create_sheet(title=subject_name[:31])

        headers = [
            "Unit",
            "Chapter",
            "Importance",
            "Week",
            "Start Date",
            "End Date",
            "Topic",
            "Estimated Hours"
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # for row in rows:
        #     ws.append([
        #         row["unit"],
        #         row["importance"],
        #         row["unit_title"],
        #         row["topic"],
        #         row["subtopics"] if row["topic"]=="SELF STUDY" else "",
        #         row["minimum_hours"] or ""
        #     ])

        self_study_written = set()

        for row in rows:
            unit = row["unit"]

            # NORMAL TOPICS
            # NORMAL TOPICS
            if not row.get("self_study", False):
                ws.append([
                    row["unit"],
                    row["unit_title"],
                    row["importance"],
                    row["week"],
                    row["start_date"],
                    row["end_date"],
                    row["topic"],
                    row["estimated_hours"] or ""
                ])
                continue


            # SELF STUDY HEADER (once per unit)
            if row["topic"] == "SELF STUDY" and unit not in self_study_written:
                # SELF STUDY ITEMS
                ws.append([
                    "", "", "",
                    row["week"],
                    row["start_date"],
                    row["end_date"],
                    row["topic"],
                    ""
                ])

                self_study_written.add(unit)
                # continue

            # SELF STUDY ITEMS (below header, Topic column only)
            ws.append([
                "", "", "",
                row.get("subtopics", ""),
                ""
            ])

            

        for col in ["E", "F"]:  # Start Date, End Date
            for cell in ws[col][1:]:
                cell.number_format = "yyyy-mm-dd"


        self._apply_unit_separators(ws)
        self._merge_units(ws)


    def _apply_unit_separators(self, ws):
        thin = Side(style="thin")
        border = Border(top=thin, bottom=thin)

# ws is a worksheet
# iter_rows() returns rows as tuples of cells
# min_row=2 means: Skip row 1 (the header)

        for row in ws.iter_rows(min_row=2):
            if row[3].value == "SELF STUDY":
                row[3].font=Font(bold=True)
                row[3].border = border
                # for cell in row:

    def save(self, path: str):
        props = self.wb.properties
        props.creator = socket.gethostname()
        props.lastModifiedBy = socket.gethostname()
        props.title = "Semester Study Planner"
        props.description = "Auto-generated semester planner using Python"
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.wb.save(path)

    def _merge_units(self, ws):
        current_unit = None
        start_row = None

        for row_idx in range(2, ws.max_row + 1):
            unit_cell = ws.cell(row=row_idx, column=1)

            if unit_cell.value != current_unit:
                if current_unit is not None:
                    ws.merge_cells(start_row=start_row, end_row=row_idx - 1,
                                start_column=1, end_column=1)
                    ws.merge_cells(start_row=start_row, end_row=row_idx - 1,
                                start_column=2, end_column=2)
                    ws.merge_cells(start_row=start_row, end_row=row_idx - 1,
                                start_column=3, end_column=3)

                current_unit = unit_cell.value
                start_row = row_idx

        # merge last unit
        if start_row:
            ws.merge_cells(start_row=start_row, end_row=ws.max_row,
                        start_column=1, end_column=1)
            ws.merge_cells(start_row=start_row, end_row=ws.max_row,
                        start_column=2, end_column=2)
            ws.merge_cells(start_row=start_row, end_row=ws.max_row,
                        start_column=3, end_column=3)

    def write_master_sheet(self, rows):
        ws = self.wb.create_sheet(title="Master")

        headers = [
            "Subject",
            "Total Units",
            "Total Topics",
            "Credits",
            "Mid %",
            "End %",
            "Total Weeks"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        subjects = {}

        for r in rows:
            subj = r["subject"]
            subjects.setdefault(subj, {
                "units": set(),
                "topics": 0,
                "weeks": set(),
                "credits": r.get("credits"),
                "mid": None,
                "end": None
            })

            subjects[subj]["units"].add(r["unit"])
            subjects[subj]["topics"] += 1
            subjects[subj]["weeks"].add(r["week"])

        for subj, data in subjects.items():
            ws.append([
                subj,
                len(data["units"]),
                data["topics"],
                data["credits"],
                data["mid"],
                data["end"],
                len(data["weeks"])
            ])
